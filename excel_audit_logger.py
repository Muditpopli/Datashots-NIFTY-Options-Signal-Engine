"""
Excel Audit Logger
Writes baseline and analysis snapshots to an Excel workbook for audit/ML use.
"""

from __future__ import annotations

from datetime import datetime
from typing import Dict, Optional
import os

import config

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None


class ExcelAuditLogger:
    BASELINE_SHEET = "baseline_snapshots"
    ANALYSIS_SHEET = "analysis_comparison"

    BASELINE_HEADERS = [
        "recorded_at",
        "trade_date",
        "index",
        "baseline_timestamp",
        "baseline_spot",
        "baseline_atm",
        "strikes_count",
        "near_atm_strikes_count",
        "atm_ce_premium",
        "atm_pe_premium",
        "atm_straddle_premium",
        "total_ce_oi_near_atm",
        "total_pe_oi_near_atm",
        "pcr_near_atm",
        "avg_ce_iv_near_atm",
        "avg_pe_iv_near_atm",
        "avg_ce_delta_near_atm",
        "avg_pe_delta_near_atm",
        "avg_ce_gamma_near_atm",
        "avg_pe_gamma_near_atm",
        "avg_ce_theta_near_atm",
        "avg_pe_theta_near_atm",
        "avg_ce_vega_near_atm",
        "avg_pe_vega_near_atm",
    ]

    ANALYSIS_HEADERS = [
        "recorded_at",
        "trade_date",
        "index",
        "signal_id",
        "baseline_timestamp",
        "analysis_timestamp",
        "baseline_spot",
        "current_spot",
        "spot_change",
        "spot_change_pct",
        "baseline_atm",
        "current_atm",
        "strikes_count",
        "near_atm_strikes_count",
        "atm_ce_premium_now",
        "atm_pe_premium_now",
        "atm_straddle_premium_now",
        "signal_direction",
        "signal_strength",
        "signal_confidence",
        "confidence_threshold",
        "passed_confidence_gate",
        "delta_flow",
        "vega_flow",
        "gamma_shift",
        "trade_type",
        "trade_direction",
        "trade_strike",
        "trade_option_type",
        "trade_edge_ratio",
        "trade_quality",
        "trade_theta_daily",
        "trade_total_risk",
        "trade_verdict",
    ]

    def __init__(self):
        self.enabled = bool(config.ENABLE_EXCEL_LOGGING)
        self.filepath = os.path.join(config.REPORTS_DIR, "dns_audit.xlsx")
        self._missing_lib_warned = False

        if self.enabled:
            self._ensure_workbook()

    def log_baseline(self, index: str, chain: Dict, baseline: Dict) -> None:
        if not self.enabled:
            return

        metrics = self._extract_chain_metrics(chain)
        row = {
            "recorded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "trade_date": datetime.now().strftime("%Y-%m-%d"),
            "index": index,
            "baseline_timestamp": baseline.get("timestamp"),
            "baseline_spot": baseline.get("spot"),
            "baseline_atm": baseline.get("atm"),
            **metrics,
        }
        self._append_row(self.BASELINE_SHEET, self.BASELINE_HEADERS, row)

    def log_analysis(
        self,
        index: str,
        chain: Dict,
        signal: Dict,
        signal_id: str,
        confidence_threshold: float,
        passed_confidence_gate: bool,
        trade: Optional[Dict] = None,
        run_phase: str = "analysis",
    ) -> None:
        if not self.enabled:
            return

        metrics = self._extract_chain_metrics(chain)
        trade_flat = self._flatten_trade(trade, run_phase=run_phase)
        row = {
            "recorded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "trade_date": datetime.now().strftime("%Y-%m-%d"),
            "index": index,
            "signal_id": signal_id,
            "baseline_timestamp": signal.get("baseline_timestamp", ""),
            "analysis_timestamp": datetime.now().isoformat(),
            "baseline_spot": signal.get("baseline_spot"),
            "current_spot": signal.get("spot"),
            "spot_change": signal.get("spot_change"),
            "spot_change_pct": signal.get("spot_change_pct"),
            "baseline_atm": signal.get("baseline_atm", ""),
            "current_atm": signal.get("atm"),
            **metrics,
            "signal_direction": signal.get("direction"),
            "signal_strength": signal.get("strength"),
            "signal_confidence": signal.get("confidence"),
            "confidence_threshold": confidence_threshold,
            "passed_confidence_gate": passed_confidence_gate,
            "delta_flow": signal.get("components", {}).get("delta_flow"),
            "vega_flow": signal.get("components", {}).get("vega_flow"),
            "gamma_shift": signal.get("components", {}).get("gamma_shift"),
            **trade_flat,
        }
        self._append_row(self.ANALYSIS_SHEET, self.ANALYSIS_HEADERS, row)

    def _ensure_workbook(self) -> None:
        if Workbook is None or load_workbook is None:
            if not self._missing_lib_warned:
                print("⚠️ Excel logging disabled: openpyxl not installed.")
                self._missing_lib_warned = True
            self.enabled = False
            return

        if os.path.exists(self.filepath):
            wb = load_workbook(self.filepath)
        else:
            wb = Workbook()
            default = wb.active
            wb.remove(default)

        if self.BASELINE_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(self.BASELINE_SHEET)
            ws.append(self.BASELINE_HEADERS)

        if self.ANALYSIS_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(self.ANALYSIS_SHEET)
            ws.append(self.ANALYSIS_HEADERS)

        wb.save(self.filepath)

    def _append_row(self, sheet_name: str, headers: list[str], row: Dict) -> None:
        if not self.enabled:
            return

        wb = load_workbook(self.filepath)
        ws = wb[sheet_name]
        ws.append([row.get(h, "") for h in headers])
        wb.save(self.filepath)

    def _extract_chain_metrics(self, chain: Dict) -> Dict:
        strikes = chain.get("strikes", [])
        spot = float(chain.get("spot", 0) or 0)
        atm = int(chain.get("atm", 0) or 0)
        strike_gap = config.STRIKE_GAPS.get(chain.get("index", "NIFTY"), 50)
        near_limit = 3 * strike_gap

        near = [s for s in strikes if abs(float(s["strike"]) - atm) <= near_limit]
        near_count = len(near)

        atm_row = next((s for s in strikes if int(s["strike"]) == atm), None)
        atm_ce = float(atm_row["ce"]["premium"]) if atm_row else 0.0
        atm_pe = float(atm_row["pe"]["premium"]) if atm_row else 0.0

        def _avg(values):
            vals = [float(v) for v in values if v is not None]
            return round(sum(vals) / len(vals), 6) if vals else 0.0

        ce_oi = sum(float(s["ce"]["oi"]) for s in near) if near else 0.0
        pe_oi = sum(float(s["pe"]["oi"]) for s in near) if near else 0.0
        pcr = (pe_oi / ce_oi) if ce_oi > 0 else 0.0

        return {
            "strikes_count": len(strikes),
            "near_atm_strikes_count": near_count,
            "atm_ce_premium": atm_ce,
            "atm_pe_premium": atm_pe,
            "atm_straddle_premium": round(atm_ce + atm_pe, 6),
            "atm_ce_premium_now": atm_ce,
            "atm_pe_premium_now": atm_pe,
            "atm_straddle_premium_now": round(atm_ce + atm_pe, 6),
            "total_ce_oi_near_atm": round(ce_oi, 2),
            "total_pe_oi_near_atm": round(pe_oi, 2),
            "pcr_near_atm": round(pcr, 6),
            "avg_ce_iv_near_atm": _avg(s["ce"]["iv"] for s in near),
            "avg_pe_iv_near_atm": _avg(s["pe"]["iv"] for s in near),
            "avg_ce_delta_near_atm": _avg(s["ce"]["delta"] for s in near),
            "avg_pe_delta_near_atm": _avg(s["pe"]["delta"] for s in near),
            "avg_ce_gamma_near_atm": _avg(s["ce"]["gamma"] for s in near),
            "avg_pe_gamma_near_atm": _avg(s["pe"]["gamma"] for s in near),
            "avg_ce_theta_near_atm": _avg(s["ce"]["theta"] for s in near),
            "avg_pe_theta_near_atm": _avg(s["pe"]["theta"] for s in near),
            "avg_ce_vega_near_atm": _avg(s["ce"]["vega"] for s in near),
            "avg_pe_vega_near_atm": _avg(s["pe"]["vega"] for s in near),
        }

    def _flatten_trade(self, trade: Optional[Dict], run_phase: str = "analysis") -> Dict:
        if not trade:
            return {
                "trade_type": "",
                "trade_direction": "",
                "trade_strike": "",
                "trade_option_type": "",
                "trade_edge_ratio": "",
                "trade_quality": "",
                "trade_theta_daily": "",
                "trade_total_risk": "",
                "trade_verdict": "READING_ONLY" if run_phase == "midday" else "NO_TRADE",
            }

        if trade.get("trade_type") == "DIRECTIONAL":
            best = trade.get("best_strike", {})
            return {
                "trade_type": "DIRECTIONAL",
                "trade_direction": trade.get("direction", ""),
                "trade_strike": best.get("strike", ""),
                "trade_option_type": best.get("option_type", ""),
                "trade_edge_ratio": best.get("edge_ratio", ""),
                "trade_quality": best.get("quality", ""),
                "trade_theta_daily": best.get("theta_daily", ""),
                "trade_total_risk": best.get("total_risk", ""),
                "trade_verdict": "TRADE",
            }

        return {
            "trade_type": trade.get("trade_type", ""),
            "trade_direction": "SIDEWAYS",
            "trade_strike": (
                f"PE {trade.get('pe_strike', '')} / CE {trade.get('ce_strike', '')}"
                if trade.get("ce_strike") or trade.get("pe_strike")
                else trade.get("strike", "")
            ),
            "trade_option_type": trade.get("setup", "DELTA_NEUTRAL"),
            "trade_edge_ratio": trade.get("edge_ratio", ""),
            "trade_quality": trade.get("quality", ""),
            "trade_theta_daily": trade.get("theta_daily", ""),
            "trade_total_risk": trade.get("total_risk", ""),
            "trade_verdict": "TRADE",
        }
