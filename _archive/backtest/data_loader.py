from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import csv
import json
import time as time_module
import requests

from .backtest_config import BacktestConfig
try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


class BacktestDataLoader:
    """
    Data layer for backtests.
    Only writes into data/backtest/* to avoid disturbing live ecosystem.
    """

    def __init__(self, cfg: Optional[BacktestConfig] = None):
        self.cfg = cfg or BacktestConfig()
        self._ensure_dirs()

    def _ensure_dirs(self) -> None:
        self.cfg.data_root.mkdir(parents=True, exist_ok=True)
        self.cfg.cache_dir.mkdir(parents=True, exist_ok=True)
        self.cfg.outputs_dir.mkdir(parents=True, exist_ok=True)

    def load_recorded_signals(
        self,
        index: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
    ) -> List[Dict]:
        if not self.cfg.predictions_csv.exists():
            return []

        rows = self._read_csv(self.cfg.predictions_csv)
        filtered = [r for r in rows if r.get("index") == index]

        if start_date:
            filtered = [r for r in filtered if str(r.get("timestamp", ""))[:10] >= start_date]
        if end_date:
            filtered = [r for r in filtered if str(r.get("timestamp", ""))[:10] <= end_date]

        return filtered

    def load_recorded_outcomes(self) -> Dict[str, Dict]:
        if not self.cfg.outcomes_csv.exists():
            return {}
        rows = self._read_csv(self.cfg.outcomes_csv)
        return {r["signal_id"]: r for r in rows if r.get("signal_id")}

    def load_analysis_audit_rows(self, index: str) -> List[Dict]:
        """
        Read analysis rows from Excel audit logger (if available).
        """
        audit_path = Path("data") / "reports" / "dns_audit.xlsx"
        if load_workbook is None or not audit_path.exists():
            return []

        wb = load_workbook(audit_path, data_only=True, read_only=True)
        if "analysis_comparison" not in wb.sheetnames:
            return []

        ws = wb["analysis_comparison"]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []

        out: List[Dict] = []
        for row in rows_iter:
            record = {}
            for i, key in enumerate(headers):
                record[key] = row[i] if i < len(row) else None
            if str(record.get("index", "")) == index:
                out.append(record)
        return out

    def write_json_output(self, filename: str, payload: Dict) -> Path:
        path = self.cfg.outputs_dir / filename
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2)
        return path

    def write_csv_output(self, filename: str, rows: List[Dict]) -> Path:
        path = self.cfg.outputs_dir / filename
        if not rows:
            path.write_text("", encoding="utf-8")
            return path
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        return path

    def fetch_dhan_historical_options_stub(
        self,
        index: str,
        from_date: str,
        to_date: str,
    ) -> None:
        """
        Placeholder hook for future Dhan rolling-option integration.
        Kept separate so live code remains untouched.
        """
        raise NotImplementedError(
            "Historical options fetch is not wired yet in this scaffold. "
            "Add a Dhan rolling-options client adapter in backtest/ only."
        )

    def fetch_dhan_rolling_options(
        self,
        index: str,
        from_date: str,
        to_date: str,
        security_id: Optional[int] = None,
        instrument: Optional[str] = None,
        strike_range: int = 10,
        interval: int = 1,
        expiry_flags: Optional[List[str]] = None,
        expiry_codes: Optional[List[int]] = None,
        sleep_seconds: float = 0.15,
        max_retries: int = 3,
    ) -> Dict:
        """
        Fetch historical rolling options from Dhan into isolated cache.
        Data is written as raw JSON responses per parameter chunk.
        """
        import config

        symbol = str(index).upper()
        default_underlyings = {
            "NIFTY": {"security_id": 13, "instrument": "OPTIDX"},
            "BANKNIFTY": {"security_id": 25, "instrument": "OPTIDX"},
        }
        merged_underlyings = {**default_underlyings, **getattr(config, "DHAN_UNDERLYING_MAP", {})}

        if security_id is None or instrument is None:
            if symbol not in merged_underlyings:
                raise ValueError(
                    f"Unsupported symbol for fetch: {index}. "
                    "Pass --security-id and --instrument, or configure DHAN_UNDERLYING_MAP_JSON."
                )
            resolved = merged_underlyings[symbol]
            sec_id = int(security_id) if security_id is not None else int(resolved["security_id"])
            instr = str(instrument).upper() if instrument is not None else str(resolved["instrument"]).upper()
        else:
            sec_id = int(security_id)
            instr = str(instrument).upper()

        token = config.DHAN_ACCESS_TOKEN.strip('"').strip("'")
        client = config.DHAN_CLIENT_ID.strip('"').strip("'")
        if not token or not client:
            raise ValueError("Missing DHAN credentials in .env")

        expiry_flags = expiry_flags or ["WEEK", "MONTH"]
        expiry_codes = expiry_codes or [1, 2, 3]

        strike_labels = ["ATM"]
        for i in range(1, strike_range + 1):
            strike_labels.append(f"ATM+{i}")
            strike_labels.append(f"ATM-{i}")

        chunks = self._date_chunks(from_date, to_date, days=30)
        endpoint = "https://api.dhan.co/v2/charts/rollingoption"
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "access-token": token,
            "client-id": client,
        }

        root = self.cfg.cache_dir / "rolling_options" / index
        root.mkdir(parents=True, exist_ok=True)

        total = 0
        success = 0
        failed = 0
        failures: List[Dict] = []

        for chunk_from, chunk_to in chunks:
            for expiry_flag in expiry_flags:
                for expiry_code in expiry_codes:
                    for strike in strike_labels:
                        for option_type in ("CALL", "PUT"):
                            total += 1
                            payload = {
                                "exchangeSegment": "NSE_FNO",
                                "interval": interval,
                                "securityId": sec_id,
                                "instrument": instr,
                                "expiryFlag": expiry_flag,
                                "expiryCode": expiry_code,
                                "strike": strike,
                                "drvOptionType": option_type,
                                "requiredData": ["close", "iv", "oi", "strike", "spot"],
                                "fromDate": chunk_from,
                                "toDate": chunk_to,
                            }

                            fname = (
                                f"{chunk_from}_{chunk_to}_{expiry_flag}_E{expiry_code}_"
                                f"{strike}_{option_type}.json"
                            ).replace(":", "_")
                            fpath = root / fname
                            if fpath.exists():
                                success += 1
                                continue

                            ok = False
                            err = ""
                            for attempt in range(max_retries):
                                try:
                                    resp = requests.post(
                                        endpoint,
                                        headers=headers,
                                        json=payload,
                                        timeout=30,
                                    )
                                    body = resp.json() if resp.content else {}
                                    status = str(body.get("status", "")).lower() if isinstance(body, dict) else ""
                                    if resp.status_code == 200 and (not status or status == "success"):
                                        with open(fpath, "w", encoding="utf-8") as f:
                                            json.dump(body, f)
                                        success += 1
                                        ok = True
                                        break

                                    err = f"http={resp.status_code} body={body}"
                                except Exception as e:
                                    err = str(e)

                                time_module.sleep(min(1.0, sleep_seconds * (attempt + 1) * 2))

                            if not ok:
                                failed += 1
                                failures.append(
                                    {
                                        "from": chunk_from,
                                        "to": chunk_to,
                                        "expiry_flag": expiry_flag,
                                        "expiry_code": expiry_code,
                                        "strike": strike,
                                        "option_type": option_type,
                                        "error": err,
                                    }
                                )

                            time_module.sleep(sleep_seconds)

        manifest = {
            "index": index,
            "security_id": sec_id,
            "instrument": instr,
            "from_date": from_date,
            "to_date": to_date,
            "strike_range": strike_range,
            "interval": interval,
            "expiry_flags": expiry_flags,
            "expiry_codes": expiry_codes,
            "total_requests": total,
            "success_requests": success,
            "failed_requests": failed,
            "generated_at": datetime.now().isoformat(),
            "cache_dir": str(root),
            "failures_preview": failures[:50],
        }
        self.write_json_output(
            f"rolling_fetch_manifest_{index}_{from_date}_{to_date}.json",
            manifest,
        )
        return manifest

    @staticmethod
    def _read_csv(path: Path) -> List[Dict]:
        with open(path, "r", newline="", encoding="utf-8") as f:
            return list(csv.DictReader(f))

    @staticmethod
    def parse_ts(value: str) -> datetime:
        return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")

    @staticmethod
    def _date_chunks(from_date: str, to_date: str, days: int = 30) -> List[tuple[str, str]]:
        start = datetime.strptime(from_date, "%Y-%m-%d").date()
        end = datetime.strptime(to_date, "%Y-%m-%d").date()
        out: List[tuple[str, str]] = []
        cur = start
        from datetime import timedelta

        if days < 1:
            days = 1

        while cur <= end:
            # Inclusive chunk window [cur, nxt], then advance to nxt+1 day.
            nxt = min(cur + timedelta(days=days - 1), end)
            out.append((cur.strftime("%Y-%m-%d"), nxt.strftime("%Y-%m-%d")))
            cur = nxt + timedelta(days=1)
        return out
